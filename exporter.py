# -*- coding: utf-8 -*-
"""
导出模块 - 生成Excel表格
"""

import os
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import logging

from config import OUTPUT_CONFIG, FIELD_MAPPING, OUTPUT_COLUMNS

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self, output_dir: str = None):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir or OUTPUT_CONFIG["output_dir"]
        self.file_prefix = OUTPUT_CONFIG["file_prefix"]
        self.datetime_format = OUTPUT_CONFIG["datetime_format"]
        self.excel_config = OUTPUT_CONFIG["excel"]
        
        # 确保输出目录存在
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def _transform_data(self, data: List[Dict]) -> pd.DataFrame:
        """
        转换数据为DataFrame
        
        Args:
            data: 原始数据列表
            
        Returns:
            pandas DataFrame
        """
        # 字段映射（英文 -> 中文）
        reverse_mapping = {v: k for k, v in FIELD_MAPPING.items()}
        
        # 构建转换后的数据
        transformed = []
        for item in data:
            row = {}
            for col_name in OUTPUT_COLUMNS:
                # 查找对应的英文字段名
                eng_name = reverse_mapping.get(col_name)
                if eng_name and eng_name in item:
                    row[col_name] = item[eng_name]
                else:
                    # 尝试直接匹配
                    for key, value in item.items():
                        if FIELD_MAPPING.get(key) == col_name:
                            row[col_name] = value
                            break
                    else:
                        row[col_name] = ""
            transformed.append(row)
        
        df = pd.DataFrame(transformed, columns=OUTPUT_COLUMNS)
        return df
    
    def _style_worksheet(self, ws, df: pd.DataFrame):
        """
        设置工作表样式
        
        Args:
            ws: openpyxl工作表
            df: 数据DataFrame
        """
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='微软雅黑', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4')
        )
        
        # 设置列宽
        column_widths = {
            '公告标题': 50,
            '发布时间': 15,
            '公告发布单位': 30,
            '项目预算': 15,
            '招标文件获取时间': 25,
            '招标报名截止时间': 25,
            '报名费用': 12,
            '投标保证金': 15,
            '项目类型': 15,
            '项目地区': 15,
            '公告类型': 12,
            '详情链接': 40,
        }
        
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            width = column_widths.get(col_name, 15)
            ws.column_dimensions[col_letter].width = width
        
        # 设置表头样式
        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置数据行样式
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # 隔行变色
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # 设置行高
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
        
        # 冻结首行
        ws.freeze_panes = self.excel_config.get("freeze_panes", "A2")
    
    def export(self, data: List[Dict], filename: str = None) -> str:
        """
        导出数据到Excel
        
        Args:
            data: 要导出的数据列表
            filename: 自定义文件名（可选）
            
        Returns:
            导出的文件路径
        """
        if not data:
            logger.warning("没有数据可导出")
            return None
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime(self.datetime_format)
            filename = f"{self.file_prefix}_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 转换数据
        df = self._transform_data(data)
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = self.excel_config.get("sheet_name", "招标公告列表")
        
        # 写入数据
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        # 应用样式
        self._style_worksheet(ws, df)
        
        # 添加汇总信息工作表
        ws_summary = wb.create_sheet(title="汇总信息")
        summary_data = [
            ["抓取时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["公告总数", len(data)],
            ["时间范围", "最近48小时"],
            ["搜索关键词", "无纸化会议"],
            ["数据来源", "乙方宝 (www.yfbzb.com)"],
        ]
        
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        # 设置汇总表样式
        for row in ws_summary.iter_rows(min_row=1, max_row=len(summary_data), min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 40
        
        # 保存文件
        wb.save(filepath)
        logger.info(f"Excel文件已保存: {filepath}")
        
        return filepath
    
    def export_csv(self, data: List[Dict], filename: str = None) -> str:
        """
        导出数据到CSV
        
        Args:
            data: 要导出的数据列表
            filename: 自定义文件名（可选）
            
        Returns:
            导出的文件路径
        """
        if not data:
            logger.warning("没有数据可导出")
            return None
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime(self.datetime_format)
            filename = f"{self.file_prefix}_{timestamp}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 转换并导出
        df = self._transform_data(data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        logger.info(f"CSV文件已保存: {filepath}")
        return filepath


def main():
    """测试函数"""
    # 测试数据
    test_data = [
        {
            "title": "某单位无纸化会议系统采购项目",
            "publish_time": "2024/12/16",
            "publish_unit": "某政府机关",
            "project_budget": "50万元",
            "bid_file_time": "2024-12-16至2024-12-20",
            "registration_deadline": "2024-12-25 17:00",
            "registration_fee": "500元",
            "bid_bond": "5000元",
            "project_type": "政府采购",
            "region": "北京",
            "announcement_type": "招标公告",
            "detail_url": "https://www.yfbzb.com/test",
        }
    ]
    
    exporter = ExcelExporter()
    filepath = exporter.export(test_data)
    print(f"测试文件已导出: {filepath}")


if __name__ == "__main__":
    main()


